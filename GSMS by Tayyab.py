import sys          # exit feature is used
import openpyxl     #Is used to write 
import pandas       # is used to read data frames
import datetime     # is used to export date and time with bills

class Login:
    def __init__ (self):
        self.obj_employee = Employee()
        self.obj_owner = Owner()
    def owner_ids(self):
        return (self.obj_owner.get_ownerids())

    def owner_login(self,id):
        return (self.obj_owner.owner_password_check(id))
    
    def employee_ids(self):
        return (self.obj_employee.get_ids())
        
    def employee_login(self,id):
        return self.obj_employee.password_check(id)


class Owner:
    def __init__ (self):
        self.id_dict = {1:'Tom', 2:'Eric'}
        self.password = {1:'tom1', 2:'eric2'}
        self.coustmer_data= [] #name & phone
        self.stock  = None
        self.data = None    #product ids validation
        self.data1 = [] #ids
        self.data2 = [] #products
        self.purchase_list = [] #items bought nested list
        self.update_data = 0   # used to subtract product from the store 
        self.sales = []
        self.bought_items=[]
        self.data_f = [] #holds nested list of whole purchase
        self.datetime = datetime.datetime.now() # to access current date and time

    def get_ownerids(self):
        return list(self.id_dict.keys())
        
    def owner_password_check(self,id):
        count = 0
        if id in (list(self.id_dict.keys())):
            owner_pass_id = input("Hey! %s Please Enter your password.: "%(self.id_dict[id])) 
            if owner_pass_id == self.password[id]:
                print ("Welcome! Mr.%s to GSMS Owner interface."%(self.id_dict[id]))
            
            while owner_pass_id != self.password[id]and count<3:
                tries = (2-count)
                print ("Mr.Owner, Com'on how can you forget this. Let's retry.\n")
                owner_pass_id = input("%s %s more chance to Re-enter your password: "%(self.id_dict[id],tries))
                count +=1
                if owner_pass_id == self.password[id]:
                    print ("Welcome to GSMS Owner interface %s"%(self.id_dict[id]))
                    return
                else:
                    if count == 3:
                        sys.exit("Next time do remember you password.")
            return     
        else:
            sys.exit("Do you really own this store")
        return

    def coustmers(self):
        column = ['Name','Phone No.']
        data = pandas.read_excel('Coustmers.xlsx',names=column, header=None,usecols=['Name','Phone No.'])
        return data
    
    def print_details(self):
        columns = ['Product_Name','Quantity','Price']
        total = 0
        bought_q = 0
        
        for i in (self.purchase_list):
            poduct_name = i[1]
            quantity = i[2]
            cost=i[3]
            total += cost
            bought_q += quantity
            data = [poduct_name,quantity,cost]
            self.data_f.append(data)
        self.data_f.append(columns)
        amount = round(total, 2) 
        data = [bought_q,amount]
        #writing amount to seprate sheet
        wb=openpyxl.load_workbook('Payments.xlsx')
        sh1=wb.active
        sh1=wb['Payments']
        sh1.append(data)
        wb.save("Payments.xlsx")
        
        data = self.data_f
        #it will take take data from the nested list and column names are stored at the end
        df = pandas.DataFrame(data[:-1],columns=data[-1])
             
        print (df)
        print ("The items you bought %s Total Due amount %s" %(bought_q,amount))

        wb=openpyxl.load_workbook('SalesRecord.xlsx')
        sh1=wb.active
        sh1=wb['SalesRecord']
        sh1.append(['_',self.datetime,'_',])
        for data in self.data_f[:-1]:
            sh1.append(data)
        wb.save("SalesRecord.xlsx")
        self.data_f=[]
        self.purchase_list = []
        return
    
    
    
    def sales_record(self):
        column = ['Product','Quantity','Price']
        record = pandas.read_excel('SalesRecord.xlsx',names=column, header=None,usecols=['Product','Quantity','Price'])
        return record

    def stock_ids(self):
        self.stock = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
        self.data1 = list(self.stock['Product Id'])
        return self.data1

    def stock_products(self):
        self.stock = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
        self.data2 = self.stock['Product Name']
        return self.data2

    def stock(self):
        self.stock = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
        self.data = list(self.stock['Product Id'])
        #it decrements the purchase quantity
        if len(self.purchase_list)>0:
            for i in range(1):
                for j in range (len(self.purchase_list)):
                    wb=openpyxl.load_workbook('InventoryData.xlsx')
                    sh1=wb.active
                    sh1=wb['InventoryData']
                    value = (self.purchase_list[j][0])+1
                    sh1.cell(row=value,column=3).value=self.purchase_list[j][-1]
                    wb.save("InventoryData.xlsx")
        self.stock = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])

        return self.stock

    def vstock(self):
        stock_data = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
        return stock_data

    def add_stock(self):
        menu_choice = int(input("1. Existing Product\n2. New Product\n> Enter Choice: "))
        if menu_choice == 1:
            self.stock1 = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
            product_id  = int(input("Enter Product ID : "))
            stock_up  = int(input("Enter Stock : "))
            new = (self.stock1.at[product_id-1,'Quantity Available'])+ stock_up
            productname =  (self.stock1.at[product_id-1,'Product Name'])
            wb=openpyxl.load_workbook('InventoryData.xlsx')
            sh1=wb.active
            sh1=wb['InventoryData']
            sh1.cell(row=product_id+1,column=3).value=new 
            wb.save("InventoryData.xlsx")
            popup =  ("%s stock has been added to the inventory new Quantity is:%s" % (productname,new))
        if menu_choice == 2:
            p_name = input("Enter Product name: ")
            stock_up = input("Enter Stock: ")
            cost_p = float(input("Enter Cost: "))
            id_product = (self.data1[-1])+1
            new_data = [id_product,p_name,stock_up,cost_p]
            wb=openpyxl.load_workbook('InventoryData.xlsx')
            sh1=wb.active
            sh1=wb['InventoryData']
            sh1.append(new_data)
            wb.save("InventoryData.xlsx")
            popup =("%s has been added to the inventory with ID:%s and Quantity:%s" % (p_name,id_product,stock_up))
            
        return popup

#inheritence
class Employee(Owner):
    
    def __init__ (self):
        super(). __init__()
        self.id_dict = {1:'Ali', 2:'Eric', 3:'Thomas', 4:'Kaif',5:'Furqan'}
        self.password = {1:'ali1', 2:'eric2', 3:'thomas3', 4:'kaif4',5:'furqan5'}
    
    def get_ids (self):
        return list(self.id_dict.keys())

    def password_check(self,id):
        if id in (list(self.id_dict.keys())):
            pass_id = input("Hey! %s Please Enter your password: "%(self.id_dict[id])) 
            if pass_id == self.password[id]:
                print ("Welcome to GSMS %s"%(self.id_dict[id]))
            count = 0
            while pass_id != self.password[id]and count<3:
                tries = (2-count)
                print ("Your password is incorrect.")
                pass_id = input("%s %s more chance to Re-enter your password: "%(self.id_dict[id],tries))
                count +=1
                if pass_id == self.password[id]:
                    print ("Welcome to GSMS %s"%(self.id_dict[id]))
                    return
                else:
                    if count == 3:
                        sys.exit("Next time do remember you password.")
            return     
        else:
            sys.exit("Your ID was not found in Database.")
        return
        
    def view_stock(self):
        return super().stock()
    
    def customer(self):
        c_name = input("Coustmer Name:")
        c_phone = input("Enter Phone:")
        data = [c_name,c_phone]
        wb=openpyxl.load_workbook('Coustmers.xlsx')
        sh1=wb.active
        sh1=wb['Coustmers']
        sh1.append(data)
        wb.save("Coustmers.xlsx")
        
        return

    def add_product(self):

        
        product_id = int(input("\nEnter Product ID: "))
        self.data = list(self.stock['Product Id'])
        
        if product_id in self.data:
            quantity = int(input("How many %s ? : "% (self.stock.at[product_id-1,'Product Name'])))
            p_name = self.stock.at[product_id-1,'Product Name']
            cost = 0

            if self.stock.at[product_id-1,'Quantity Available'] < 7 and self.stock.at[product_id-1,'Quantity Available'] > 0:
                print ( "%s stock is running low only %s items left."%(self.stock.at[product_id-1,'Product Name'],(self.stock.at[product_id-1,'Quantity Available'])-quantity))
            if self.stock.at[product_id-1,'Quantity Available'] < quantity and self.stock.at[product_id-1,'Quantity Available'] > 0:
                print ( "Available But in Less Quantity.")
                if quantity > self.stock.at[product_id-1,'Quantity Available']:
                    quantity = self.stock.at[product_id-1,'Quantity Available']
                    cost = quantity* (self.stock.at[product_id-1,'Cost'])
                    self.update_data = self.stock.at[product_id-1,'Quantity Available'] - quantity
                    purchase = [product_id,p_name,quantity,cost,self.update_data]
                    self.purchase_list.append(purchase)
                
            if self.stock.at[product_id-1,'Quantity Available'] <= 0:
                print ("Product Not available.")
                
            if self.stock.at[product_id-1,'Quantity Available'] > quantity:
                cost = quantity* (self.stock.at[product_id-1,'Cost'])
                self.update_data = self.stock.at[product_id-1,'Quantity Available'] - quantity
                purchase = [product_id,p_name,quantity,cost,self.update_data]
                self.purchase_list.append(purchase)

                

                
        else:
            print ("product Id not matched")
        return
   
    def get_Receipt(self):
        column = ['Quantity','Price']
        record = pandas.read_excel('Payments.xlsx',names=column, header=None,usecols=['Quantity','Price'])
        price = list(record['Price'])
        quantity = list(record['Quantity'])
        p = price[-1]
        q = quantity[-1]
        return ("Total Quantity = %s Due Amount = %s\n"% (q,p))

    def e_pay(self):
        amount = float(input("Enter Amount to pay: "))
        column = ['Quantity','Price']
        record = pandas.read_excel('Payments.xlsx',names=column, header=None,usecols=['Quantity','Price'])
        price = list(record['Price'])
        p = price[-1]
        if amount>p:
            total = round((amount-p),2)
            print ("\nThanks for paying your remaining amount is: %s "% (total))
        if amount<p:
            total1 = round((p-amount),2)
            print ("\nThanks for paying your but you still have to pay : %s "% (total1))
        if amount==p:
            print ("\nThanks for paying!")
        return
    
    def remove_product(self):

        print ("\nID.|Product Name")
        for i in range(1):
            for j in range(len(self.purchase_list)):
                print ("%s. %s "% (j,self.purchase_list[j][1]))
        list_index = int(input("Enter the product No. you want to remove:"))
        del_list = self.purchase_list[list_index]

        if len(self.purchase_list)>0:
            for i in range(1):
                for j in range (len(del_list)):
                    wb=openpyxl.load_workbook('InventoryData.xlsx')
                    sh1=wb.active
                    sh1=wb['InventoryData']
                    value = (del_list[0])+1
                    p_value = (del_list[2])+(del_list[-1])
                            
                    sh1.cell(row=value,column=3).value=p_value
                    wb.save("InventoryData.xlsx")
        
        del self.purchase_list[list_index]
        self.stock = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
        return "Removed Successfully."
    def printing(self):
        return super().print_details() #inheritence
    

#composition  
class Customer:
    def __init__(self):
        self.obj_employee = Employee()
    def getReceipt(self):
        return(self.obj_employee.get_Receipt())
    def pay(self):
        return self.obj_employee.e_pay()   
    
#aggregation
class Product_stock:

    def __init__(self,employ):
        self.employ = employ
        
    def product_store(self):
        return self.employ.view_stock()
    def update_product(self):
        menu_choice = int(input("1.Refill Existing Product\n2.Quit\n> Enter Choice: "))
        if menu_choice == 2:
            return ('No item was updated')
        if menu_choice == 1:
            self.stock1 = pandas.read_excel('InventoryData.xlsx',usecols=['Product Id', 'Product Name','Quantity Available','Cost'])
            product_id  = int(input("Enter Product ID : "))
            stock_up  = int(input("Enter Stock : "))
            new = (self.stock1.at[product_id-1,'Quantity Available'])+ stock_up
            productname =  (self.stock1.at[product_id-1,'Product Name'])
            wb=openpyxl.load_workbook('InventoryData.xlsx')
            sh1=wb.active
            sh1=wb['InventoryData']
            sh1.cell(row=product_id+1,column=3).value=new 
            wb.save("InventoryData.xlsx")
            popup1 =  ("%s stock has been Refilled new Quantity is:%s" % (productname,new))
        return popup1

def main():
    if __name__ ==  "__main__":

        
        while True:
            print("\t\tGSMS LogIn.\n1.Owner\n2.Employee\n3.Quit\n")
            choice = int(input("\nPlease choose one of the above: "))
            verfiction = Login()
            employ = Employee()
            data = Employee()
            owner= Owner()
            customer = Customer()
            productstore = Product_stock(employ)
            employ.view_stock()
            
            if choice == 3:
                break
            if choice == 2:
                print ("\nThe registered Employees ID are: %s " % (verfiction.employee_ids()))
                id = int(input("\nPlease Enter your ID: "))
                verfiction.employee_login(id)
                while True:
                    emp_choice= int(input("\n1. View Stock\n2. New Coustmer\n3. LogOut \n> Make a choice: "))
                    if emp_choice == 1:
                        print (employ.view_stock())
                    if emp_choice==2:
                        employ.customer()
                        choice = int(input("\n1. Add Product\n2. Quit\n>Enter choice :  "))  
                        if choice == 1:
                            employ.add_product()
                            while True:
                                choice = int(input("\n1. Add Product\n2. Remove product \n3. Billing Counter\n>Enter choice :  "))
                                if choice == 1:
                                    employ.view_stock()
                                    employ.add_product()
                                    employ.view_stock()
                                if choice == 2:
                                    employ.view_stock()
                                    print(employ.remove_product())
                                    employ.view_stock()
                                if choice == 3:
                                    employ.view_stock()
                                    employ.printing()
                                    employ.view_stock()
                                    print("\nHere is your Short Receipt:")
                                    print (customer.getReceipt())
                                    customer.pay()   
                                    break
                        
                        
                        while True:
                            menup= int(input("\nStock Popup! \n1. Recheck\n2. Refil\n3. Exit\n>Enter choice :"))
                            if menup == 1:
                                print(productstore.product_store())
                            if menup == 2:
                                print(productstore.update_product())
                            if menup == 3:
                                break      
                    if emp_choice == 3:
                        break
   
            if choice == 1:
                print ("\nThe registered Owner ID are: %s " % (verfiction.owner_ids()))
                id = int(input("\nPlease Enter your ID: "))
                verfiction.owner_login(id)
                owner.stock_ids()
        
                
                person = Owner()
                while True:
                    onr_choice= int(input("\n1. View Stock\n2. View sales \n3. Coustmers List\n4. Update Stock\n5. LogOut \n> Make a choice: "))
                    if onr_choice == 1:
                        print (owner.vstock())
                    if onr_choice == 2:
                        print (owner.sales_record())
                    if onr_choice==3:
                        print (owner.coustmers())
                    if onr_choice==4:
                        print (owner.add_stock())
                    if onr_choice == 5:
                        break
    
            
            
main()
